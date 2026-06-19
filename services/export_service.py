import os
import csv
import logging

logger = logging.getLogger(__name__)


# ─── CSV ──────────────────────────────────────────────────────────────────────

def export_csv(filepath, data, headers):
    """Exports a list of dicts to a UTF-8 CSV file."""
    try:
        with open(filepath, mode='w', newline='', encoding='utf-8-sig') as f:
            writer = csv.DictWriter(f, fieldnames=headers, extrasaction='ignore')
            writer.writeheader()
            writer.writerows(data)
        return filepath
    except Exception as e:
        logger.error(f"CSV export error: {e}")
        return None


# ─── Excel ────────────────────────────────────────────────────────────────────

def export_excel(filepath, data, headers, title="Report", meta=None):
    """
    Exports to a styled Excel workbook.
    meta: optional dict of {label: value} shown at top (company name, period, etc.)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                     numbers)
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Sheet name max 31 chars

        # ── Colour palette ──
        HEADER_FILL  = PatternFill("solid", fgColor="1C2434")   # dark blue
        HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        TITLE_FONT   = Font(bold=True, name="Calibri", size=14, color="101828")
        META_FONT    = Font(name="Calibri", size=10, color="667085")
        ALT_FILL     = PatternFill("solid", fgColor="F9FAFB")
        BORDER_SIDE  = Side(style="thin", color="CED4DA")
        CELL_BORDER  = Border(bottom=Border(bottom=BORDER_SIDE).bottom)
        CENTER       = Alignment(horizontal="center", vertical="center")
        LEFT         = Alignment(horizontal="left",   vertical="center")
        RIGHT        = Alignment(horizontal="right",  vertical="center")

        row = 1

        # ── Title ──
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=max(len(headers), 2))
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = TITLE_FONT
        cell.alignment = LEFT
        ws.row_dimensions[row].height = 28
        row += 1

        # ── Meta block ──
        if meta:
            for label, value in meta.items():
                ws.cell(row=row, column=1, value=label).font = META_FONT
                ws.cell(row=row, column=2, value=str(value)).font = META_FONT
                row += 1
        ws.cell(row=row, column=1, value=f"Generated: {__import__('datetime').datetime.now().strftime('%d %b %Y %H:%M')}").font = META_FONT
        row += 2  # blank spacer

        # ── Column headers ──
        for col_idx, header in enumerate(headers, 1):
            c = ws.cell(row=row, column=col_idx, value=header)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = Border(
                top=Side(style="thin", color="000000"),
                bottom=Side(style="thin", color="000000"),
            )
        ws.row_dimensions[row].height = 22
        header_row = row
        row += 1

        # ── Data rows ──
        for data_idx, row_data in enumerate(data):
            fill = ALT_FILL if data_idx % 2 == 0 else PatternFill()
            for col_idx, key in enumerate(headers, 1):
                value = row_data.get(key, "")
                c = ws.cell(row=row, column=col_idx, value=value)
                c.fill = fill
                c.alignment = RIGHT if col_idx == len(headers) else LEFT
                c.font = Font(name="Calibri", size=10)
                c.border = Border(bottom=Side(style="hair", color="DEE2E6"))
            ws.row_dimensions[row].height = 18
            row += 1

        # ── Auto column widths ──
        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(str(header))
            for row_data in data:
                val = str(row_data.get(header, ""))
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

        # ── Freeze pane under header ──
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # ── Auto-filter on headers ──
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(headers))}{header_row}"
        )

        wb.save(filepath)
        return filepath

    except ImportError:
        logger.error("openpyxl is not installed.")
        return None
    except Exception as e:
        logger.error(f"Excel export error: {e}")
        return None


# ─── PDF ──────────────────────────────────────────────────────────────────────

def export_pdf(filepath, title, data, headers, meta=None):
    """
    Exports to a professional branded PDF using reportlab.
    meta: optional dict of {label: value} shown in the header block.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable)
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

        # ── Color palette ──
        DARK_BLUE  = colors.HexColor("#1C2434")
        LIGHT_GREY = colors.HexColor("#F9FAFB")
        MID_GREY   = colors.HexColor("#667085")
        BORDER_CLR = colors.HexColor("#DEE2E6")
        GREEN      = colors.HexColor("#12B76A")
        RED        = colors.HexColor("#F04438")
        WHITE      = colors.white

        doc = SimpleDocTemplate(
            filepath, pagesize=A4,
            leftMargin=20*mm, rightMargin=20*mm,
            topMargin=18*mm, bottomMargin=18*mm
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Normal"],
            fontSize=18, leading=24, textColor=DARK_BLUE,
            fontName="Helvetica-Bold", spaceAfter=4
        )
        meta_style = ParagraphStyle(
            "Meta", parent=styles["Normal"],
            fontSize=9, textColor=MID_GREY, fontName="Helvetica"
        )
        cell_style = ParagraphStyle(
            "Cell", parent=styles["Normal"],
            fontSize=9, fontName="Helvetica", leading=12
        )

        elements = []

        # ── Header block ──
        elements.append(Paragraph(title, title_style))
        elements.append(HRFlowable(width="100%", thickness=2, color=DARK_BLUE, spaceAfter=6))

        # Meta info (period, company, generated date)
        now_str = __import__('datetime').datetime.now().strftime('%d %b %Y %H:%M')
        if meta:
            for label, value in meta.items():
                elements.append(Paragraph(f"<b>{label}:</b> {value}", meta_style))
        elements.append(Paragraph(f"Generated: {now_str}", meta_style))
        elements.append(Spacer(1, 10))

        # ── Table ──
        # Build table data
        t_data = [[Paragraph(f"<b>{h}</b>", ParagraphStyle(
            "H", parent=cell_style, textColor=WHITE, fontName="Helvetica-Bold"
        )) for h in headers]]

        for row in data:
            t_data.append([
                Paragraph(str(row.get(h, "")), cell_style) for h in headers
            ])

        # Dynamic column widths
        page_width = A4[0] - 40*mm
        col_widths = [page_width / len(headers)] * len(headers)
        # Last column right-aligned (usually Amount)
        if len(headers) > 1:
            col_widths[-1] = 40*mm
            remaining = page_width - 40*mm
            col_widths[:-1] = [remaining / (len(headers)-1)] * (len(headers)-1)

        table = Table(t_data, colWidths=col_widths, repeatRows=1)

        row_count = len(t_data)
        style_cmds = [
            # Header row
            ("BACKGROUND",    (0, 0), (-1, 0), DARK_BLUE),
            ("TEXTCOLOR",     (0, 0), (-1, 0), WHITE),
            ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE",      (0, 0), (-1, 0), 9),
            ("ROWBACKGROUND", (0, 0), (-1, 0), DARK_BLUE),
            # Data rows
            ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE",      (0, 1), (-1, -1), 9),
            ("TEXTCOLOR",     (0, 1), (-1, -1), DARK_BLUE),
            # Alternating rows
            *[("BACKGROUND", (0, i), (-1, i), LIGHT_GREY)
              for i in range(2, row_count, 2)],
            # Grid
            ("GRID",          (0, 0), (-1, -1), 0.5, BORDER_CLR),
            ("ROWHEIGHT",     (0, 0), (-1, -1), 18),
            # Alignment
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN",         (-1, 0), (-1, -1), "RIGHT"),  # last col right
            ("ALIGN",         (0, 0), (-2, -1), "LEFT"),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
        ]

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

        # ── Footer ──
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=BORDER_CLR))
        footer_style = ParagraphStyle(
            "Footer", parent=styles["Normal"],
            fontSize=8, textColor=MID_GREY,
            fontName="Helvetica", alignment=TA_RIGHT
        )
        elements.append(Paragraph("KTIB CashFlow — Confidential Financial Report", footer_style))

        doc.build(elements)
        return filepath

    except ImportError:
        logger.error("reportlab is not installed.")
        return None
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        return None


# ─── JSON Full Backup ──────────────────────────────────────────────────────────

def export_all_json(filepath, company_id):
    """Full data backup to JSON (accounts, transactions, categories, planned, budgets, projects)."""
    import json
    from database.session import get_session
    from database.models import Account, Transaction, Category, PlannedPayment, Budget, Project

    def dt(d):
        return d.isoformat() if d else None

    try:
        with get_session() as session:
            accounts = [
                {"id": a.id, "name": a.name, "type": a.type, "currency": a.currency,
                 "balance": float(a.balance), "color": a.color,
                 "identifier": a.identifier, "is_archived": a.is_archived}
                for a in session.query(Account).filter_by(company_id=company_id).all()
            ]
            transactions = [
                {"id": t.id, "account_id": t.account_id, "to_account_id": t.to_account_id,
                 "category_id": t.category_id, "project_id": t.project_id,
                 "type": t.type, "amount": float(t.amount), "currency": t.currency,
                 "description": t.description, "counterparty": t.counterparty,
                 "date": dt(t.date), "note": t.note, "is_recurring": t.is_recurring,
                 "recurring_type": t.recurring_type, "status": t.status,
                 "edv_amount": float(t.edv_amount) if t.edv_amount else 0.0,
                 "base_amount": float(t.base_amount) if t.base_amount else None,
                 "base_edv_amount": float(t.base_edv_amount) if t.base_edv_amount else None}
                for t in session.query(Transaction).filter_by(company_id=company_id).all()
            ]
            categories = [
                {"id": c.id, "parent_id": c.parent_id, "name": c.name,
                 "type": c.type, "color": c.color, "icon": c.icon, "is_default": c.is_default}
                for c in session.query(Category).filter_by(company_id=company_id).all()
            ]
            planned = [
                {"id": p.id, "account_id": p.account_id, "category_id": p.category_id,
                 "project_id": p.project_id,
                 "type": p.type, "amount": float(p.amount), "currency": p.currency,
                 "edv_amount": float(p.edv_amount) if p.edv_amount else 0.0,
                 "description": p.description, "counterparty": p.counterparty,
                 "due_date": dt(p.due_date), "status": p.status,
                 "recurring": p.recurring, "next_due_date": dt(p.next_due_date),
                 "paid_at": dt(p.paid_at)}
                for p in session.query(PlannedPayment).filter_by(company_id=company_id).all()
            ]
            budgets = [
                {"id": b.id, "category_id": b.category_id, "amount": float(b.amount),
                 "period_type": b.period_type, "month": b.month, "year": b.year}
                for b in session.query(Budget).filter_by(company_id=company_id).all()
            ]
            projects = [
                {"id": p.id, "name": p.name, "description": p.description,
                 "color": p.color,
                 "budget": float(p.budget) if p.budget else None,
                 "start_date": dt(p.start_date), "end_date": dt(p.end_date),
                 "status": p.status}
                for p in session.query(Project).filter_by(company_id=company_id).all()
            ]

            payload = {
                "export_date": dt(__import__('datetime').datetime.now()),
                "company_id": company_id,
                "accounts": accounts,
                "transactions": transactions,
                "categories": categories,
                "planned_payments": planned,
                "budgets": budgets,
                "projects": projects,
            }

        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False)

        return filepath
    except Exception as e:
        logger.error(f"JSON export error: {e}")
        return None


def clear_all_data(company_id):
    """Deletes all transactions, planned payments, categories, accounts and projects for a company."""
    from database.session import get_session
    from database.models import Account, Transaction, Category, PlannedPayment, Project

    with get_session() as session:
        session.query(Transaction).filter_by(company_id=company_id).delete()
        session.query(PlannedPayment).filter_by(company_id=company_id).delete()
        session.query(Category).filter_by(company_id=company_id).delete()
        session.query(Account).filter_by(company_id=company_id).delete()
        session.query(Project).filter_by(company_id=company_id).delete()
        session.commit()
